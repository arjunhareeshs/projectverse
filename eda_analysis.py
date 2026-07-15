import openpyxl
from collections import Counter
import statistics

wb = openpyxl.load_workbook("data/Final Groups and Project Registration.xlsx", data_only=True)

# I Year, II Year Students wise D (3626 rows)
print("="*80)
print("I Year, II Year Students wise D (3626 rows) - EDA")
print("="*80)
ws = wb['I Year, II Year Students wise D']
years = []
depts = []
clusters = []
genders = []
residents = []
modes = []
ssg = []
group_reg = []
skills_reg = []
domains = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] is not None: years.append(str(row[3]).strip())
    if row[4] is not None: depts.append(str(row[4]).strip())
    if row[6] is not None: clusters.append(str(row[6]).strip())
    if row[7] is not None: genders.append(str(row[7]).strip())
    if row[8] is not None: residents.append(str(row[8]).strip())
    if row[9] is not None: modes.append(str(row[9]).strip())
    if row[10] is not None: ssg.append(str(row[10]).strip())
    if row[11] is not None: group_reg.append(str(row[11]).strip())
    if row[12] is not None: skills_reg.append(str(row[12]).strip())
    if row[13] is not None: domains.append(str(row[13]).strip())
print("Year:", Counter(years).most_common())
print("Cluster:", Counter(clusters).most_common())
print("Gender:", Counter(genders).most_common())
print("Resident:", Counter(residents).most_common())
print("Learning Mode:", Counter(modes).most_common())
print("SSG:", Counter(ssg).most_common())
print("Group Registered:", Counter(group_reg).most_common())
print("Skills Registered:", Counter(skills_reg).most_common())
print("SSG Domain:", Counter(domains).most_common())
print("Top 5 depts:", Counter(depts).most_common(5))

print()
print("="*80)
print("GROUP (sample) (1000 rows) - EDA")
print("="*80)
ws = wb['GROUP (sample)']
team_sizes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    count = sum(1 for c in row[4:20] if c is not None and str(c).strip() != "")
    team_sizes.append(count)
print("Total groups:", len(team_sizes))
print("Team size stats: min=", min(team_sizes), "max=", max(team_sizes), "avg=", sum(team_sizes)/len(team_sizes))
size_dist = Counter(team_sizes)
print("Team size distribution:", sorted(size_dist.items()))

print()
print("="*80)
print("INDUSTRY MENTORS (264 rows)")
print("="*80)
ws = wb['Industry mentors']
mentor_filled = 0
mentor_industries = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[9] is not None and str(row[9]).strip() != "":
        mentor_filled += 1
        mentor_industries.append(str(row[9]).strip())
print("Groups with industry mentor name:", mentor_filled)
print("Groups without industry mentor:", ws.max_row - 1 - mentor_filled)
print("Top mentor industries:", Counter(mentor_industries).most_common(10))

print()
print("="*80)
print("PEER TO PEER REVIEW SCHEDULE (229 rows) - EDA")
print("="*80)
ws = wb['Peer to Peer Review Schedule']
venues = []
times = []
years_p = []
groups_p = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] is not None: years_p.append(str(row[3]).strip())
    if row[4] is not None: groups_p.append(str(row[4]).strip())
    if row[5] is not None: times.append(str(row[5]).strip())
    if row[6] is not None: venues.append(str(row[6]).strip())
print("Year:", Counter(years_p).most_common())
print("Unique groups in P2P:", len(set(groups_p)))
print("Time slots:", Counter(times).most_common())
print("Venues:", Counter(venues).most_common())

print()
print("="*80)
print("USERS (8715 rows)")
print("="*80)
ws = wb['users']
ids = []
emails = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is not None: ids.append(row[0])
    if row[1] is not None: emails.append(str(row[1]).strip())
print("Total user records:", len(ids))
print("Unique emails:", len(set(emails)))

print()
print("="*80)
print("USER SKILL WISE RANK (19355 rows)")
print("="*80)
ws = wb['User Skill wise rank']
skill_ids = []
skills = []
points = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is not None: skill_ids.append(row[0])
    if row[1] is not None: skills.append(str(row[1]).strip())
    if row[5] is not None:
        try: points.append(float(row[5]))
        except: pass
print("Total skill-user records:", len(skills))
print("Unique skills:", len(set(skills)))
print("Points stats: min=", min(points), "max=", max(points), "avg=", sum(points)/len(points))
print("Top 10 skills by frequency:", Counter(skills).most_common(10))

print()
print("="*80)
print("USER SKILL MAPPING (19389 rows)")
print("="*80)
ws = wb['User Skill Mapping']
types = []
total_points = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[6] is not None: types.append(str(row[6]).strip())
    if row[2] is not None:
        try: total_points.append(float(row[2]))
        except: pass
print("Type distribution:", Counter(types).most_common())
print("Total points: min=", min(total_points), "max=", max(total_points), "avg=", sum(total_points)/len(total_points))

print()
print("="*80)
print("PROJECT SKILLS SET MAPPING (4083 rows)")
print("="*80)
ws = wb['Project Skills Set Mapping']
skills_p = []
projects = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is not None: projects.append(row[0])
    if row[1] is not None: skills_p.append(str(row[1]).strip())
print("Total project-skill links:", len(skills_p))
print("Unique projects:", len(set(projects)))
print("Top 10 skills:", Counter(skills_p).most_common(10))

print()
print("="*80)
print("GROUP RANKING (1000 rows)")
print("="*80)
ws = wb['Group Ranking']
points = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4] is not None:
        try: points.append(float(row[4]))
        except: pass
print("Total rankings:", len(points))
print("Points stats: min=", min(points), "max=", max(points), "avg=", sum(points)/len(points))

print()
print("="*80)
print("SSG GROUP REGISTRATION PENDING (778 rows)")
print("="*80)
ws = wb['SSG Group Registration Pending']
years_p = []
depts_p = []
genders_p = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[3] is not None: years_p.append(str(row[3]).strip())
    if row[4] is not None: depts_p.append(str(row[4]).strip())
    if row[7] is not None: genders_p.append(str(row[7]).strip())
print("Total Hardware SSG pending:", len(years_p))
print("Year:", Counter(years_p).most_common())
print("Gender:", Counter(genders_p).most_common())
print("Top depts:", Counter(depts_p).most_common(5))

print()
print("="*80)
print("Copy of Group Registration (3358 rows) - EDA")
print("="*80)
ws = wb['Copy of Group Registration']
years = []
roles = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] is not None: years.append(str(row[3]).strip())
    if row[5] is not None: roles.append(str(row[5]).strip())
print("Year:", Counter(years).most_common())
print("Roles:", Counter(roles).most_common(20))

print()
print("="*80)
print("Sheet42 (1000 rows) - EDA")
print("="*80)
ws = wb['Sheet42']
years = []
clusters = []
genders = []
modes = []
ssgs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[5] is not None: years.append(str(row[5]).strip())
    if row[8] is not None: clusters.append(str(row[8]).strip())
    if row[9] is not None: genders.append(str(row[9]).strip())
    if row[11] is not None: modes.append(str(row[11]).strip())
    if row[13] is not None: ssgs.append(str(row[13]).strip())
print("Year:", Counter(years).most_common())
print("Cluster:", Counter(clusters).most_common())
print("Gender:", Counter(genders).most_common())
print("Mode:", Counter(modes).most_common())
print("SSG:", Counter(ssgs).most_common())
