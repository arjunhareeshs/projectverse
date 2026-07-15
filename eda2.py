import openpyxl
from collections import Counter
import statistics

wb = openpyxl.load_workbook("data/Final Groups and Project Registration.xlsx", data_only=True)

# Group Summary - more details
print("="*80)
print("GROUP SUMMARY (264 rows) - EDA")
print("="*80)
ws = wb['Group Summary']
levels = []
total_count = []
iyear_count = []
iiyear_count = []
max_limits = []
avail = []
proj_reg = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] is not None: levels.append(row[1])
    if row[8] is not None:
        try: total_count.append(float(row[8]))
        except: pass
    if row[6] is not None:
        try: iyear_count.append(float(row[6]))
        except: pass
    if row[7] is not None:
        try: iiyear_count.append(float(row[7]))
        except: pass
    if row[9] is not None: max_limits.append(row[9])
    if row[10] is not None: avail.append(row[10])
    if row[11] is not None: proj_reg.append(str(row[11]).strip())
print("Group Levels:", Counter(levels).most_common())
print("Total Count stats: min=", min(total_count), "max=", max(total_count), "avg=", sum(total_count)/len(total_count))
print("I Year Count stats: min=", min(iyear_count), "max=", max(iyear_count), "sum=", sum(iyear_count))
print("II Year Count stats: min=", min(iiyear_count), "max=", max(iiyear_count), "sum=", sum(iiyear_count))
print("Max Members Limits:", Counter(max_limits).most_common())
print("Project Registered:", Counter(proj_reg).most_common())
print("Total groups:", ws.max_row - 1)
print("Groups with total >=9:", sum(1 for x in total_count if x >= 9))

print()
print("="*80)
print("DEPARTMENT SUMMARY - Cluster wise (left side) - EDA")
print("="*80)
ws = wb['Department Summary']
# Rows 2 has column labels for the clusters
# Cluster: cols 2(Non CS), 3(CS), 4(TOTAL)
# Sample first 5 groups cluster count
for r in range(3, 10):
    grp = ws.cell(r, 1).value
    noncs = ws.cell(r, 2).value
    cs = ws.cell(r, 3).value
    tot = ws.cell(r, 4).value
    print(f"  Group {grp}: NonCS={noncs}, CS={cs}, Total={tot}")

# Department wise (cols 6-17)
print()
print("Sample Department wise counts (rows 3-10):")
for r in range(3, 10):
    grp = ws.cell(r, 1).value
    ag = ws.cell(r, 6).value
    aids = ws.cell(r, 7).value
    aiml = ws.cell(r, 8).value
    cse = ws.cell(r, 10).value
    ece = ws.cell(r, 12).value
    it = ws.cell(r, 14).value
    print(f"  Group {grp}: AG={ag}, AIDS={aids}, AIML={aiml}, CSE={cse}, ECE={ece}, IT={it}")

print()
print("="*80)
print("PROJECT REGISTRATIONS - skill count per project")
print("="*80)
ws = wb['Project Registrations']
skill_counts = []
for row in ws.iter_rows(min_row=2, values_only=True):
    count = sum(1 for c in row[8:33] if c is not None and str(c).strip() != "")
    skill_counts.append(count)
print("Project-skill links: min=", min(skill_counts), "max=", max(skill_counts), "avg=", sum(skill_counts)/len(skill_counts))
print("Distribution:", Counter(skill_counts).most_common())

# Get unique project titles for category check
print()
print("="*80)
print("PROJECT REGISTRATIONS - unique groups with projects")
print("="*80)
ws = wb['Project Registrations']
groups = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2] is not None: groups.append(str(row[2]).strip())
print("Unique groups with projects:", len(set(groups)))
print("Top groups by project count:", Counter(groups).most_common(5))
print("Total project entries:", len(groups))
